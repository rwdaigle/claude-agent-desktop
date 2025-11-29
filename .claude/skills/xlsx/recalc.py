# /// script
# requires-python = ">=3.12"
# dependencies = ["xlcalculator>=0.5.0", "openpyxl>=3.1.0"]
# ///

"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using xlcalculator (pure Python).

Note: xlcalculator supports most common Excel functions but may not support
all advanced functions. Unsupported formulas will show errors.
"""

import json
import sys
from pathlib import Path

from openpyxl import load_workbook
from xlcalculator import ModelCompiler, Evaluator


def recalc(filename, timeout=30):
    """
    Recalculate formulas in Excel file and report any errors

    Args:
        filename: Path to Excel file
        timeout: Maximum time to wait for recalculation (seconds) - not used in pure Python approach

    Returns:
        dict with error locations and counts
    """
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    try:
        # Compile the workbook model and evaluate formulas
        compiler = ModelCompiler()
        model = compiler.read_and_parse_archive(filename)
        evaluator = Evaluator(model)
        evaluator.evaluate()

        # Load workbook to write calculated values back
        wb = load_workbook(filename)

        # Track which cells have formulas vs calculated values
        formula_count = 0
        cells_updated = 0

        # Update cells with calculated values
        for cell_address, cell_model in model.cells.items():
            if cell_model.formula:
                formula_count += 1
                try:
                    # Parse sheet name and coordinate from address
                    if "!" in cell_address:
                        sheet_name, coord = cell_address.split("!")
                        # Remove quotes from sheet name if present
                        sheet_name = sheet_name.strip("'")
                    else:
                        # Default to first sheet if no sheet specified
                        sheet_name = wb.sheetnames[0]
                        coord = cell_address

                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        ws[coord].value = cell_model.value
                        cells_updated += 1
                except Exception:
                    # Skip cells that can't be updated
                    pass

        wb.save(filename)
        wb.close()

        # Check for Excel errors in the recalculated file
        wb = load_workbook(filename, data_only=True)

        excel_errors = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        # Build result summary
        result = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "total_formulas": formula_count,
            "cells_updated": cells_updated,
            "error_summary": {},
        }

        # Add non-empty error categories
        for err_type, locations in error_details.items():
            if locations:
                result["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],  # Show up to 20 locations
                }

        return result

    except Exception as e:
        return {"error": str(e)}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file using xlcalculator")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - cells_updated: Number of cells with recalculated values")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        print("\nNote: xlcalculator supports most common Excel functions but may")
        print("not support all advanced functions. Unsupported formulas will show errors.")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
